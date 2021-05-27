from openpyxl import Workbook, load_workbook
from datetime import datetime


# 添加联合表单
wb = load_workbook('courses.xlsx')
ws = wb[wb.sheetnames[0]]
wt = wb[wb.sheetnames[1]]
wc = wb.create_sheet(title='combine')

wc.append(['创建时间', '课程名称', '学习人数', '学习时间'])
for i in list(ws.values)[1:]:
    for j in wt.values:
        if i[1] == j[1]:
            wc.append(list(i)+[j[-1]])

wb.save('courses.xlsx')


# 按年份创建新的表单
d = {i[0].strftime('%Y'): Workbook() for i in wc.values if type(i[0]) == datetime}

values = wc.values
next(values)

for i in values:
    wb = d[i[0].strftime('%Y')]
    wb.active.append(i)

for year, wb in d.items():
    wb.active.title = year
    wb.save('{}.xlsx'.format(year))
