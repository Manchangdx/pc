from openpyxl import Workbook, load_workbook

wb = load_workbook('courses.xlsx')
ws = wb[wb.sheetnames[0]]
wt = wb[wb.sheetnames[1]]
wc = wb.create_sheet(title='combine')

wc.append(['创建时间', '课程名称', '学习人数', '学习时间'])
for a, b in zip(list(ws.values)[1:], list(wt.values)[1:]):
        wc.append(list(a)+[b[-1]])
wb.save('courses.xlsx')

s = set()
l = list(wc.values)[1:]
for i in l:
    s.add(i[0].strftime('%Y'))
for y in s:
    wb = Workbook()
    ws = wb.active
    ws.title = y
    for i in l:
        if i[0].strftime('%Y') == y:
            ws.append(i)
    wb.save('{}.xlsx'.format(y))
